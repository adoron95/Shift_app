import json
from datetime import timedelta

import openpyxl

from request import sorce as req, guard
from order import sorce as shift
from sort import sorce as srt

# 
def open_jsonfile():
    account = {}
    # Opening JSON file
    f = open('gurd.json')

    # returns JSON object as
    # a dictionary
    data = json.load(f)
    # Iterating through the json
    # list
    for name in data:
        account[name] = 0
        account[name] += len(data[name]["full_morning"])
        account[name] += len(data[name]["half_morning"])
    return account
    # Closing file
    f.close()


def initial_guards_data():
    path = "guards data.xlsx"
    wb_write = openpyxl.load_workbook(path)
    sw = wb_write.active
    guards_dict = {}
    guards_list = []
    max_r = sw.max_row
    max_col = sw.max_column

    for name in sw.iter_cols(0, max_col):
        for i in range(3, max_r):
            if name[i].column == 1:
                guards_dict[name[i].value] = guard.Guard(name[i])
                guards_dict[name[i].value].name=name[i].value
                guards_list.append(name[i].value)
                continue
            if name[i].column == 2 and name[i].value is not None:
                guards_dict[guards_list[i - 3]].manager = True
                continue

            if name[i].column == 3 and name[i].value is not None:
                guards_dict[guards_list[i - 3]].achsan = True
                continue

            if name[i].column == 4 and name[i].value is not None:
                guards_dict[guards_list[i - 3]].armed = True
                continue

            if name[i].column == 5 and name[i].value is not None:
                guards_dict[guards_list[i - 3]].controller = True
                continue

            if name[i].column == 6 and name[i].value is not None:
                guards_dict[guards_list[i - 3]].religion = True
                continue

            if name[i].column == 7:
                guards_dict[guards_list[i - 3]].seniority = name[i].value
                continue
    return guards_dict


"""   data = {
        "phone": "00",
        "quality": 100,
        "armed": "False",
        "controller": "False",
        "religion": "False",
        "manager": "False",
        "full_morning": [],
        "continuity": [],
        "half_morning": [],
        "afternoon": [],
        "half_afternoon": [],
        "night": [],
        "friday": "[] Todo new file",
        "weekend1": "False",
        "weekend2": "False",
        "weekend3": "False",
        "weekend4": "False",
        "driver": "false",
        "seniority": 4.2022
    }
    with open("gurd.json", 'r+') as file:
        # First we load existing data into a dict.
        file_data = json.load(file)
        for i in list_name:
            file_data[i] = data
        # Sets file's current position at offset.
        file.seek(0)
        # convert back to json.
        json.dump(file_data, file, indent=4)"""

if __name__ == "__main__":

    #dict_guard = initial_guards_data()
    quality, raw = req.init(initial_guards_data())
    two_weeks=shift.init()
    srt.init(raw,quality,two_weeks)
